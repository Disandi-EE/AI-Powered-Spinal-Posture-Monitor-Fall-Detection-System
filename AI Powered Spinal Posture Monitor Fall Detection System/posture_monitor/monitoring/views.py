from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from .models import PostureData, PostureSession, UserProfile, EmergencyAlert
from .ml_models import posture_analyzer
import json
import csv
import io
from datetime import datetime, timedelta
from django.utils import timezone

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'home.html')

def register_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        emergency_contact = request.POST.get('emergency_contact', '')
        
        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            UserProfile.objects.create(user=user, emergency_contact=emergency_contact)
            login(request, user)
            return redirect('dashboard')
        except Exception as e:
            return render(request, 'register.html', {'error': str(e)})
    
    return render(request, 'register.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            return render(request, 'login.html', {'error': 'Invalid credentials'})
    
    return render(request, 'login.html')

@login_required
def logout_view(request):
    logout(request)
    return redirect('home')

@login_required
def dashboard(request):
    # Get user's profile
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Get recent posture data
    recent_data = PostureData.objects.filter(user=request.user)[:100]
    
    # Get today's statistics
    today = timezone.now().date()
    today_data = PostureData.objects.filter(
        user=request.user,
        timestamp__date=today
    )
    
    # Calculate statistics
    total_samples = today_data.count()
    correct_samples = today_data.filter(is_correct_posture=1).count()
    correctness_percentage = (correct_samples / total_samples * 100) if total_samples > 0 else 0
    
    context = {
        'profile': profile,
        'total_samples': total_samples,
        'correctness_percentage': round(correctness_percentage, 2),
        'recent_alerts': EmergencyAlert.objects.filter(user=request.user)[:5]
    }
    
    return render(request, 'dashboard.html', context)

@login_required
def real_time_monitoring(request):
    """
    Real-time monitoring page view with updated context for WebSocket fix
    """
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Prepare user data for JavaScript (safe template rendering)
    user_data = {
        'id': request.user.id,
        'username': request.user.username,
        'is_authenticated': True,
        'email': request.user.email,
    }
    
    context = {
        'user': request.user,
        'user_data': user_data,
        'profile': profile,
        'page_title': 'Real-time Monitoring',
    }
    
    return render(request, 'real_time_monitoring.html', context)

@login_required
def offline_analysis(request):
    return render(request, 'offline_analysis.html')

def read_csv_file(file_content):
    """Parse CSV content and return list of dictionaries"""
    # Decode file content if it's bytes
    if isinstance(file_content, bytes):
        content = file_content.decode('utf-8')
    else:
        content = file_content
    
    # Use StringIO to create a file-like object
    csv_file = io.StringIO(content)
    reader = csv.DictReader(csv_file)
    
    data = []
    for row in reader:
        data.append(dict(row))
    
    return data

def read_excel_file(uploaded_file):
    """
    Basic Excel reading without openpyxl/xlrd
    Note: This is a simplified version. For full Excel support, you'd need openpyxl
    """
    try:
        # Try to import openpyxl if available
        from openpyxl import load_workbook
        
        workbook = load_workbook(uploaded_file)
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        # Get data rows
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=1):
            if any(cell is not None for cell in row):  # Skip empty rows
                row_data = {}
                for i, header in enumerate(headers):
                    if i < len(row) and row[i] is not None:
                        row_data[header] = row[i]
                data.append(row_data)
        
        return data
        
    except ImportError:
        # Fallback: suggest user to convert to CSV
        raise Exception("Excel files require openpyxl library. Please convert your file to CSV format.")

def validate_numeric_value(value, field_name):
    """Validate and convert a value to float"""
    try:
        return float(value)
    except (ValueError, TypeError):
        raise ValueError(f"Invalid numeric value for {field_name}: {value}")

@login_required
@csrf_exempt
def upload_offline_data(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        
        try:
            # Read the file based on extension
            if uploaded_file.name.endswith('.csv'):
                # Read CSV file
                file_content = uploaded_file.read()
                data = read_csv_file(file_content)
                
            elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                # Read Excel file
                data = read_excel_file(uploaded_file)
                
            else:
                return JsonResponse({'error': 'Unsupported file format. Please use CSV or Excel files.'}, status=400)
            
            # Validate required columns
            required_columns = ['tilt_x', 'tilt_y']
            if not data:
                return JsonResponse({'error': 'File is empty or has no data rows'}, status=400)
            
            # Check if required columns exist in first row
            first_row = data[0]
            missing_columns = [col for col in required_columns if col not in first_row]
            if missing_columns:
                return JsonResponse({
                    'error': f'Missing required columns: {", ".join(missing_columns)}. Available columns: {", ".join(first_row.keys())}'
                }, status=400)
            
            # Prepare and validate data for analysis
            data_list = []
            for i, row in enumerate(data):
                try:
                    tilt_x = validate_numeric_value(row.get('tilt_x'), 'tilt_x')
                    tilt_y = validate_numeric_value(row.get('tilt_y'), 'tilt_y')
                    
                    data_list.append({
                        'tilt_x': tilt_x,
                        'tilt_y': tilt_y
                    })
                except ValueError as e:
                    return JsonResponse({
                        'error': f'Data validation error at row {i + 2}: {str(e)}'
                    }, status=400)
            
            # Analyze the data
            analysis_result = posture_analyzer.analyze_batch_data(data_list)
            
            return JsonResponse({
                'success': True,
                'analysis': analysis_result,
                'rows_processed': len(data_list)
            })
            
        except Exception as e:
            return JsonResponse({'error': f'File processing error: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Invalid request. Please upload a CSV or Excel file.'}, status=400)

@login_required
def settings(request):
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        emergency_contact = request.POST.get('emergency_contact')
        profile.emergency_contact = emergency_contact
        profile.save()
        return redirect('settings')
    
    return render(request, 'settings.html', {'profile': profile})

@login_required
def api_posture_history(request):
    """API endpoint to get posture history data"""
    days = int(request.GET.get('days', 7))
    start_date = timezone.now() - timedelta(days=days)
    
    data = PostureData.objects.filter(
        user=request.user,
        timestamp__gte=start_date
    ).values('timestamp', 'is_correct_posture', 'tilt_x', 'tilt_y')
    
    return JsonResponse(list(data), safe=False)

@login_required
def get_user_data(request):
    """
    API endpoint to get current user data for WebSocket connection
    """
    user_data = {
        'id': request.user.id,
        'username': request.user.username,
        'is_authenticated': request.user.is_authenticated,
        'email': request.user.email,
    }
    return JsonResponse(user_data)

# Additional helper functions for data processing without pandas

def calculate_statistics(data_list, field):
    """Calculate basic statistics for a numeric field"""
    if not data_list:
        return {'count': 0, 'mean': 0, 'min': 0, 'max': 0}
    
    values = [float(item[field]) for item in data_list if field in item and item[field] is not None]
    
    if not values:
        return {'count': 0, 'mean': 0, 'min': 0, 'max': 0}
    
    return {
        'count': len(values),
        'mean': sum(values) / len(values),
        'min': min(values),
        'max': max(values)
    }

def group_by_date(posture_data):
    """Group posture data by date without pandas"""
    grouped = {}
    
    for data_point in posture_data:
        date_key = data_point['timestamp'].date() if hasattr(data_point['timestamp'], 'date') else data_point['timestamp']
        
        if date_key not in grouped:
            grouped[date_key] = []
        
        grouped[date_key].append(data_point)
    
    return grouped

def filter_data_by_criteria(data_list, criteria):
    """Filter data based on criteria without pandas"""
    filtered_data = []
    
    for item in data_list:
        match = 1
        for key, value in criteria.items():
            if key not in item or item[key] != value:
                match = 0
                break
        
        if match:
            filtered_data.append(item)
    
    return filtered_data